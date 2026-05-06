"""
Excel text extraction using openpyxl.

Public API
----------
extract_text_from_excel(file_bytes: bytes) -> str
"""

import io
import logging
from typing import Any, Optional

import openpyxl
from openpyxl.cell.cell import MergedCell
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)

_MAX_ROWS = 5000    # safety cap per sheet
_MAX_COLS = 200     # safety cap per sheet


def _cell_value(cell: Any) -> str:
    """Return a clean string representation of a cell value."""
    if isinstance(cell, MergedCell):
        return ""
    val = cell.value
    if val is None:
        return ""
    if isinstance(val, float):
        # Avoid spurious ".0" on integers stored as floats
        if val == int(val):
            return str(int(val))
        return str(round(val, 6))
    return str(val).strip().replace("\n", " ")


def _sheet_to_text(ws: Worksheet, sheet_name: str) -> str:
    """
    Convert a worksheet to a human-readable tab-separated text block.

    Only rows that contain at least one non-empty cell are included.
    """
    lines: list[str] = [f"Sheet: {sheet_name}"]
    lines.append("=" * 60)

    rows_written = 0
    for row_idx, row in enumerate(ws.iter_rows(), start=1):
        if row_idx > _MAX_ROWS:
            lines.append(f"[truncated after {_MAX_ROWS} rows]")
            break

        cells = [_cell_value(cell) for cell in row[: _MAX_COLS]]
        # Skip entirely empty rows
        if not any(cells):
            continue

        lines.append("\t".join(cells).rstrip())
        rows_written += 1

    if rows_written == 0:
        lines.append("[empty sheet]")

    return "\n".join(lines)


def extract_text_from_excel(file_bytes: bytes) -> str:
    """
    Extract all text from an Excel workbook (xlsx / xls via openpyxl).

    Each sheet becomes its own labelled text block.  Sheets are joined by
    ``\\n\\n``.

    Parameters
    ----------
    file_bytes:
        Raw bytes of the Excel file.

    Returns
    -------
    str
        Extracted text, or empty string on failure.
    """
    try:
        wb = openpyxl.load_workbook(
            io.BytesIO(file_bytes),
            read_only=True,        # memory-efficient for large files
            data_only=True,        # return computed values, not formulas
            keep_links=False,
        )
    except Exception as exc:
        logger.error("Failed to open Excel workbook: %s", exc)
        return ""

    sheet_texts: list[str] = []

    try:
        for sheet_name in wb.sheetnames:
            try:
                ws = wb[sheet_name]
                sheet_texts.append(_sheet_to_text(ws, sheet_name))
            except Exception as exc:
                logger.warning("Could not read sheet '%s': %s", sheet_name, exc)
                sheet_texts.append(f"Sheet: {sheet_name}\n[error reading sheet: {exc}]")
    finally:
        wb.close()

    if not sheet_texts:
        return ""

    return "\n\n".join(sheet_texts)
